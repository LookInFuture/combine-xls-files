import random

import xlsxwriter
from openpyxl import load_workbook
import os
from collections import OrderedDict

FULL_PUBLIC_LIST = OrderedDict()
MISSING_PHONE = 0
OUT_FILE = 'Full public list.xlsx'

path = os.path.join(os.path.abspath('CF'))
list_of_files = os.scandir(path)


class FileHandler:
    """
    Class FileHandler handles everything that should be done in order to prepare "Data Base" to save.
    Args:
        worksheet: Active worksheet from the opened Excel.
    Params:
        self.file = active worksheet;
        self.all_names: OrderedDict = saves all names, phones and other details in the dict;
        self.name: str = First name, Surname;
        self.email: str = Email;
        self.phone: str or tuple = Phone #
        self.source: str = Where the contact came from (Facebook, YouTube etc.)
    """
    def __init__(self, worksheet) -> None:
        self.file = worksheet
        self.all_names: OrderedDict = OrderedDict()
        self.name: str = ''
        self.email: str = ''
        self.phone: str = ''
        self.source: str = 'Unknown'

    def phone_startswith(self) -> None:
        """
        Checks if the number starts from '+' symbol if so, deletes it. As well if the phone starts from 27, replace
        27 to 0. Applied only for South African numbers.
        :return: None
        """
        if isinstance(self.phone, list):
            self.phone = [number[1:] if number.startswith('+') else number for number in self.phone]
            self.phone = [f'0{number[2:]}' if number.startswith('27') else number for number in self.phone]
        else:
            self.phone = self.phone[1:] if self.phone.startswith('+') else self.phone
            self.phone = f'0{self.phone[2:]}' if self.phone.startswith('27') else self.phone

    def phone_more_then_one(self) -> None:
        """
        Checks if there's more than one phone number in the Excell cell and if there's special symbol,
        splits all numbers and save it as a list.
        :return: None
        """
        if len(self.phone) > 10 and '/' in self.phone:
            self.phone = self.phone.split('/')
        if len(self.phone) > 10 and '\n' in self.phone:
            self.phone = self.phone.split('\n')

    def phone_remove_spaces(self) -> None:
        """
        Checks if there's spaces in the phone number, if so - deletes it.
        :return: None
        """
        if ' ' in self.phone:
            self.phone = self.phone.replace(' ', '')

    def phone_exists_in_db(self) -> None:
        """
        Checks if there's phone number in current Data Base (self.all_names). If not - creates record and updates
        main Dict (FULL_PUBLIC_LIST) where saved all public from the other excels.
        :return:
        """
        add_to_db = dict()

        try:
            if self.phone == '':
                self.phone = str(random.randint(1, 300))
            current_name = self.all_names.get(self.phone)
            if not current_name:
                check_name_in_db = self.name_exists_in_db()
                if not check_name_in_db:
                    add_to_db[self.phone] = [{
                        'name': self.name,
                        'email': self.email,
                        'source': self.source}
                    ]
                    self.all_names.update(add_to_db)

        except TypeError:
            add_to_db[tuple(self.phone)] = [{
                'name': self.name,
                'email': self.email,
                'source': self.source
            }]
            self.all_names.update(add_to_db)

        finally:
            FULL_PUBLIC_LIST.update(self.all_names)

    def name_exists_in_db(self) -> bool:
        """
        Checks if the name in main Dict (FULL_PUBLIC_LIST). If there's name, then updates email if it's empty.
        :return: Bool
        """
        for i_name in FULL_PUBLIC_LIST.values():
            try:
                if self.name in i_name[0].get('name') and not i_name[0].get('email'):
                    i_name[0].update(
                        {
                            'email': self.email,
                        }
                    )
                    return True
            except KeyError:
                continue
        return False

    def get_worksheet_details(self) -> None:
        """
        Getting all details for a row in the worksheet.
        :return: None
        """
        for item in range(1, self.file.max_row + 1):
            first_name = self.file[f'A{item}'].value or ''
            surname = self.file[f'B{str(item)}'].value or ''
            phone = self.file[f'C{str(item)}'].value or ''
            email = self.file[f'D{str(item)}'].value or ''
            source = self.file[f'E{str(item).upper().strip()}'].value or 'Div 6'
            self.name = f'{first_name.upper().strip()} {surname.upper().strip()}' if first_name else f'{surname.upper().strip()}'
            self.phone = str(phone)
            self.email = email.lower().strip()
            self.source = source.upper().strip()
            if not first_name == 'NAME':
                self.phone_remove_spaces()
                self.phone_more_then_one()
                self.phone_startswith()
                self.phone_exists_in_db()


def save_to_file() -> None:
    """
    This function works on saving all needed details that has been saved in the main Dict (FULL_PUBLIC_LIST).
    :return: None
    """
    with xlsxwriter.Workbook(OUT_FILE) as workbook:
        ws = workbook.add_worksheet()
        headers = ['NAME', 'PHONE', 'EMAIL', 'SOURCE']
        bold = workbook.add_format({'bold': True})
        for col, header in enumerate(headers):
            ws.write_string(0, col, header, cell_format=bold)

        for index, person in enumerate(FULL_PUBLIC_LIST.items(), start=1):
            if not len(person[1]) > 1:
                name = person[1][0].get('name')
                phone = str(person[0]) if isinstance(person[0], str) \
                    else str(person[0]).replace('(', '').replace(')', '').replace('\'', '')
                email = person[1][0].get('email')
                source = person[1][0].get('source')
                ws.write_string(index, 0, name)
                ws.write_string(index, 1, phone)
                ws.write_string(index, 2, email)
                ws.write_string(index, 3, source)
            else:
                for i_name in range(len(person[1])):
                    name = person[1][i_name].get('name')
                    phone = str(person[0]) if isinstance(person[0], str) \
                        else str(person[0]).replace('(', '').replace(')', '').replace('\'', '')
                    email = person[1][0].get('email')
                    source = person[1][0].get('source')
                    ws.write_string(index, 0, name)
                    ws.write_string(index, 1, phone)
                    ws.write_string(index, 2, email)
                    ws.write_string(index, 3, source)


def get_file_details(filename: str) -> None:
    """
    This function loads Excel file and getting worksheet details and initialize FileHandler class.
    :param filename: passed filename from the directory.
    :return: None
    """
    book = load_workbook(filename=filename)
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        current_file = FileHandler(sheet)
        current_file.get_worksheet_details()
        FULL_PUBLIC_LIST.update(current_file.all_names)


if __name__ == '__main__':
    for file in list_of_files:
        print('Working on following file: {current_file}'.format(current_file=file))
        get_file_details(file)
    save_to_file()
